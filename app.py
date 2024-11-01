from flask import Flask, render_template, redirect, url_for, session, request, send_file
import sqlite3
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'edu7701')


usuarios = {
    'Mod': '7701'
}

DATABASE_URL = os.getenv('DATABASE_URL', 'empresa_3d.db')

def get_db_connection():
    conn = sqlite3.connect(DATABASE_URL)
    conn.row_factory = sqlite3.Row
    return conn

def crear_tablas():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS productos (
                    id INTEGER PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    stock INTEGER NOT NULL,
                    precio REAL NOT NULL
                    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS compras (
                    id INTEGER PRIMARY KEY,
                    producto_id INTEGER,
                    cantidad INTEGER,
                    fecha TEXT,
                    FOREIGN KEY (producto_id) REFERENCES productos (id)
                    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS ventas (
                    id INTEGER PRIMARY KEY,
                    producto_id INTEGER,
                    cantidad INTEGER,
                    fecha TEXT,
                    FOREIGN KEY (producto_id) REFERENCES productos (id)
                    )''')
                    

    conn.commit()
    conn.close()

crear_tablas()


@app.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos')
    productos = cursor.fetchall()
    conn.close()
    return render_template('index.html', productos=productos)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if username in usuarios and usuarios[username] == password:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            print(f"Intento fallido de login: Usuario: {username}, Contraseña: {password}")
            return render_template('login.html')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.before_request
def require_login():

    if not session.get('logged_in') and request.endpoint != 'login':
        return redirect(url_for('login'))

@app.route('/add_product', methods=['POST'])
def add_product():
    nombre = request.form['nombre']
    stock = int(request.form['stock'])
    precio = float(request.form['precio'])

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, stock FROM productos WHERE nombre = ?', (nombre,))
    producto = cursor.fetchone()

    if producto:
        nuevo_stock = producto[1] + stock
        cursor.execute('UPDATE productos SET stock = ?, precio = ? WHERE id = ?', (nuevo_stock, precio, producto[0]))
    else:
        cursor.execute('INSERT INTO productos (nombre, stock, precio) VALUES (?, ?, ?)', (nombre, stock, precio))
    
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/registrar_compra', methods=['POST'])
def registrar_compra():
    producto_id = int(request.form['producto_id'])
    cantidad = int(request.form['cantidad'])
    fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO compras (producto_id, cantidad, fecha) VALUES (?, ?, ?)', (producto_id, cantidad, fecha))
    cursor.execute('UPDATE productos SET stock = stock + ? WHERE id = ?', (cantidad, producto_id))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/registrar_venta', methods=['POST'])
def registrar_venta():
    producto_id = int(request.form['producto_id'])
    cantidad = int(request.form['cantidad'])

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT stock FROM productos WHERE id = ?', (producto_id,))
    stock_actual = cursor.fetchone()

    if stock_actual and stock_actual[0] >= cantidad:
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('INSERT INTO ventas (producto_id, cantidad, fecha) VALUES (?, ?, ?)', (producto_id, cantidad, fecha))
        cursor.execute('UPDATE productos SET stock = stock - ? WHERE id = ?', (cantidad, producto_id))
        conn.commit()
    else:
        print(f"Error: No hay suficiente stock para realizar la venta. Stock disponible: {stock_actual[0] if stock_actual else 'Producto no encontrado'}")

    conn.close()
    return redirect(url_for('index'))

@app.route('/eliminar_producto', methods=['POST'])
def eliminar_producto():
    producto_id = int(request.form['producto_id'])

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM productos WHERE id = ?', (producto_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/mostrar_inventario')
def mostrar_inventario():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos')
    productos = cursor.fetchall()
    conn.close()
    return render_template('inventario.html', productos=productos)

@app.route('/informe')
def informe():
    c = canvas.Canvas("informe_HZ_movimientos.pdf", pagesize=A4)
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(300, 800, "Informe Compras y Ventas - HZ Impresiones 3D")
    c.setFont("Helvetica", 10)
    c.drawCentredString(300, 780, f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    y = 750
    c.drawString(100, y, "Compras realizadas:")
    y -= 20
    c.drawString(100, y, "ID|            Producto           | Cantidad | Fecha")
    y -= 20

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''SELECT compras.id, productos.nombre, compras.cantidad, compras.fecha 
                      FROM compras 
                      JOIN productos ON compras.producto_id = productos.id''')
    compras = cursor.fetchall()
    for compra in compras:
        c.drawString(100, y, f"{compra[0]} | {compra[1]} | {compra[2]} | {compra[3]}")
        y -= 20

    y -= 40
    c.drawString(100, y, "Ventas")
    y -= 20
    c.drawString(100, y, "ID|            Producto           | Cantidad | Fecha ")
    y -= 20

    cursor.execute('''SELECT ventas.id, productos.nombre, ventas.cantidad, ventas.fecha 
                      FROM ventas 
                      JOIN productos ON ventas.producto_id = productos.id''')
    ventas = cursor.fetchall()
    for venta in ventas:
        c.drawString(100, y, f"{venta[0]} | {venta[1]} | {venta[2]} | {venta[3]}")
        y -= 20
    c.save()

    return send_file("informe_HZ_movimientos.pdf", as_attachment=True)

@app.route('/reporte_excel')
def reporte_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventario"
    
    sheet.append([""])
    sheet.append(["INVENTARIO - HZ IMPRESIONES 3D"])
    sheet.append([""])
    
    sheet.merge_cells('A2:D2')
    title_cell = sheet['A2']
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.font = Font(bold=True, size=14)

    encabezados = ["ID", "Nombre", "Stock", "Precio (Q)"]
    sheet.append(encabezados)

    for col in range(1, len(encabezados) + 1):
        sheet.cell(row=4, column=col).font = Font(bold=True)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos')
    productos = cursor.fetchall()

    for producto in productos:
        sheet.append([producto[0], producto[1], producto[2], producto[3]])

    for row in sheet.iter_rows(min_row=5, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00'

    excel_file = 'inventario_HZ_impresiones.xlsx'
    workbook.save(excel_file)

    return send_file(excel_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
